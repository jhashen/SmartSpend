from flask import Flask, render_template, request, url_for, session, redirect, abort
import os
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from datetime import datetime
from io import BytesIO
from model.voucher_prediction_model import train_and_predict

app = Flask(__name__, static_folder='static')
app.secret_key = 'your_secret_key_here'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB upload limit

def verify_template_exists(template_name):
    """Verify that a template exists before rendering"""
    template_path = os.path.join(app.template_folder, template_name)
    if not os.path.exists(template_path):
        app.logger.error(f"Template not found: {template_name}")
        abort(500, description=f"Required template '{template_name}' not found in templates directory")

@app.route('/')
def home():
    verify_template_exists('frontend_design.html')
    return render_template('frontend_design.html')

@app.route('/upload', methods=['POST'])
def upload():
    verify_template_exists('frontend_design.html')
    
    if 'file' not in request.files:
        return render_template('frontend_design.html', 
                            prediction="❌ No file uploaded",
                            error=True)

    file = request.files['file']
    if file.filename == '':
        return render_template('frontend_design.html',
                            prediction="❌ Empty file uploaded",
                            error=True)

    try:
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return render_template('frontend_design.html',
                                prediction="❌ Only Excel files are supported",
                                error=True)
        
        file_bytes = file.read()
        session['file_bytes'] = file_bytes
        session['filename'] = file.filename
        
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        sheets_info = {}
        total_tables = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            tables = list(sheet.tables.keys())
            has_data = sheet.max_row > 1
            table_count = len(tables)
            sheets_info[sheet_name] = {
                'tables': tables,
                'has_data': has_data,
                'table_count': table_count
            }
            total_tables += table_count
        
        if total_tables == 1:
            for sheet_name, info in sheets_info.items():
                if info['tables']:
                    session['selected_sheet'] = sheet_name
                    session['selected_table'] = info['tables'][0]
                    return redirect(url_for('predict'))
        
        # Verify select_table exists before rendering
        verify_template_exists('select_table.html')
        return render_template('select_table.html', 
                            sheets_info=sheets_info,
                            filename=file.filename,
                            total_tables=total_tables)
    
    except Exception as e:
        return render_template('frontend_design.html',
                            prediction=f"⚠️ File Error: {str(e)}",
                            error=True)

@app.route('/predict', methods=['POST', 'GET'])
def predict():
    verify_template_exists('frontend_design.html')
    
    if 'file_bytes' not in session:
        return redirect(url_for('home'))
    
    if request.method == 'GET':
        if 'selected_sheet' not in session or 'selected_table' not in session:
            return redirect(url_for('home'))
    
    try:
        if request.method == 'POST':
            selected_sheet = request.form['sheet']
            selected_table = request.form.get('table', None)
        else:
            selected_sheet = session['selected_sheet']
            selected_table = session['selected_table']
        
        file_bytes = session['file_bytes']
        
        session.pop('file_bytes', None)
        session.pop('filename', None)
        session.pop('selected_sheet', None)
        session.pop('selected_table', None)
        
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        if selected_sheet not in wb.sheetnames:
            raise ValueError("Selected sheet does not exist")
            
        sheet = wb[selected_sheet]
        if selected_table and selected_table not in sheet.tables:
            raise ValueError("Selected table does not exist")
        
        if selected_table:
            table = sheet.tables[selected_table]
            data = sheet[table.ref]
            rows = [[cell.value for cell in row] for row in data]
            
            if len(rows) < 2:
                raise ValueError("Selected table contains no data")
                
            df = pd.DataFrame(rows[1:], columns=rows[0])
        else:
            df = pd.read_excel(
                BytesIO(file_bytes),
                sheet_name=selected_sheet,
                engine='openpyxl',
                dtype={'Amount': 'float32', 'Year': 'int16'},
                usecols="A:D",
                nrows=100000
            )
            if df.empty:
                raise ValueError("Selected sheet contains no data")
        
        result = train_and_predict(df)
        historical = pd.DataFrame(result['historical_data'])

        plt.figure(figsize=(12, 6))
        plt.plot(historical['Year'], historical['TotalSpent'], 
                marker='o', markersize=8, linewidth=2,
                label='Historical Spending', color='#1f77b4')
        plt.plot(result['future_years'], result['future_predictions'], 
               marker='o', markersize=8, linestyle='--', linewidth=2,
               label='Recommended Budget', color='#2ca02c')
        plt.fill_between(
            result['future_years'],
            [x - result['mae'] for x in result['future_predictions']],
            [x + result['mae'] for x in result['future_predictions']],
            color='#2ca02c', alpha=0.1, 
            label=f'±RM{result["mae"]:,.2f} Margin'
        )
        
        plt.title("Voucher Budget Planning", fontsize=14, pad=20)
        plt.xlabel("Year", fontsize=12)
        plt.ylabel("Total Amount (RM)", fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend(fontsize=10)
        
        for x, y in zip(historical['Year'], historical['TotalSpent']):
            plt.annotate(f'RM{y:,.0f}', (x, y), textcoords="offset points",
                        xytext=(0,10), ha='center', fontsize=9)
        
        for x, y in zip(result['future_years'], result['future_predictions']):
            plt.annotate(f'RM{y:,.0f}', (x, y), textcoords="offset points",
                        xytext=(0,10), ha='center', fontsize=9)
        
        plt.tight_layout()
        
        plot_filename = f"budget_plot_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
        plot_path = os.path.join('static', plot_filename)
        plt.savefig(plot_path, dpi=100, bbox_inches='tight')
        plt.close()

        recommendation = [
            f"<tr><td>{year}</td><td>RM {amount:,.2f}</td><td>±RM {result['mae']:,.2f}</td></tr>"
            for year, amount in zip(result['future_years'], result['future_predictions'])
        ]
        
        history_table = [
            f"<tr><td>{row['Year']}</td><td>RM {row['TotalSpent']:,.2f}</td>"
            f"<td>{row['TransactionCount']}</td><td>RM {row['AverageSpent']:,.2f}</td></tr>"
            for _, row in historical.iterrows()
        ]
        
        return render_template(
            'frontend_design.html',
            prediction={
                'recommendation': "\n".join(recommendation),
                'last_year': result['last_year'],
                'last_year_spent': f"RM {result['last_year_spent']:,.2f}",
                'growth_rate': f"{result['growth_rate']*100:.1f}%",
                'accuracy': f"±RM {result['mae']:,.2f}",
                'mape': f"{result['mape']*100:.1f}%",
                'model': result['model_type'],
                'plot_url': url_for('static', filename=plot_filename),
                'history_table': "\n".join(history_table)
            },
            success=True
        )

    except ValueError as e:
        return render_template('frontend_design.html',
                            prediction=f"⚠️ Data Error: {str(e)}",
                            error=True)
    except Exception as e:
        return render_template('frontend_design.html',
                            prediction=f"⚠️ System Error: {str(e)}",
                            error=True)

if __name__ == '__main__':
    # Verify all required templates exist before starting
    required_templates = ['frontend_design.html', 'select_table.html']
    for template in required_templates:
        if not os.path.exists(os.path.join(app.template_folder, template)):
            raise FileNotFoundError(f"Critical template missing: {template}")
    
    app.run(debug=True, host='0.0.0.0', port=5000)